import os
import mysql.connector
from mysql.connector import Error
from openpyxl import Workbook, load_workbook

def export_login_to_excel(xlsx_path="login data.xlsx"):
    try:
        mydb = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="demo"
        )
    except Error as e:
        print(f"Could not connect to MySQL: {e}")
        return

    try:
        mycursor = mydb.cursor()
        mycursor.execute("SELECT * FROM login")
        rows = mycursor.fetchall()

        if not rows:
            print("No rows returned from database.")
            return

        # Prepare workbook: load if exists, otherwise create new and add header
        new_workbook = False
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "sheet1"
            new_workbook = True

        # Add header row if workbook was just created and cursor.description is available
        if new_workbook and mycursor.description:
            headers = [col[0] for col in mycursor.description]
            ws.append(headers)

        # Append all rows
        for row in rows:
            print(row)
            ws.append(row)

        wb.save(xlsx_path)
        print(f"Data written to '{xlsx_path}' successfully")

    except Error as e:
        print(f"MySQL error: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")
    finally:
        try:
            mycursor.close()
        except Exception:
            pass
        try:
            mydb.close()
        except Exception:
            pass

def print_records_to_excel(xlsx_path="login data.xlsx"):
    """Print/display all records from sheet1 in a formatted way."""
    try:
        if not os.path.exists(xlsx_path):
            print(f"File '{xlsx_path}' does not exist.")
            return
        
        wb = load_workbook(xlsx_path)
        ws = wb["sheet1"]
        
        print(f"\n{'='*80}")
        print(f"Records from {xlsx_path} - sheet1")
        print(f"{'='*80}\n")
        
        # Get all rows
        all_rows = list(ws.iter_rows(values_only=True))
        
        if not all_rows:
            print("No records found in sheet1.")
            return
        
        # Print header
        header = all_rows[0]
        if header:
            print(" | ".join(str(h) for h in header))
            print("-" * 80)
        
        # Print data rows
        for row in all_rows[1:]:
            print(" | ".join(str(cell) if cell is not None else "" for cell in row))
        
        print(f"\n{'='*80}")
        print(f"Total records: {len(all_rows) - 1}")
        print(f"{'='*80}\n")
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")


if __name__ == "__main__":
    export_login_to_excel()
    print_records_to_excel()


    